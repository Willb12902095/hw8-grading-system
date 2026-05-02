#!/usr/bin/env python3
"""
Student Homework Submission Grader

This script:
1. Unzips student submissions from a ZIP file
2. Detects student submission folders
3. Checks for required files (code files, chat history)
4. Reads and analyzes code and chat files (safe - no execution)
5. Applies a grading rubric with high/low performance criteria
6. Creates an Excel file with grading results
7. Creates a separate rubric sheet
8. Creates a grading log text file

Uses pandas and openpyxl for Excel operations.
"""

import os
import sys
import zipfile
import re
import shutil
import subprocess
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Tuple, Optional
import pandas as pd

# Try to import openpyxl, install if needed
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
except ImportError:
    print("Installing required packages...")
    os.system("pip install openpyxl pandas")
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows


# =============================================================================
# CONFIGURATION
# =============================================================================

# Required file patterns (can be customized)
REQUIRED_CODE_EXTENSIONS = ['.py', '.java', '.js', '.ts', '.cpp', '.c', '.go', '.rs']

CHAT_HISTORY_PATTERNS = ['chat.txt', 'conversation.txt', 'history.txt', '*.txt']

# Runtime smoke-test settings. This is optional and disabled by default.
# It opens each submitted Python file briefly to check whether it launches without immediate crash.
# It does NOT judge gameplay quality and does NOT replace manual review.
RUNTIME_SMOKE_TEST_TIMEOUT_SECONDS = 5

# Grading rubric criteria
RUBRIC = {
    'code_submission': {
        'name': 'Code Submission',
        'high': 'Code files present and well-structured',
        'low': 'No code files or very minimal',
        'max_points': 20
    },
    'chat_history': {
        'name': 'Chat History',
        'high': 'Chat history present with good explanation',
        'low': 'No chat history or very brief',
        'max_points': 15
    },
    'completeness': {
        'name': 'Completeness',
        'high': 'All required components present',
        'low': 'Missing major components',
        'max_points': 20
    },
    'correctness': {
        'name': 'Code Correctness',
        'high': 'Code appears correct with good logic',
        'low': 'Code has obvious errors or issues',
        'max_points': 25
    },
    'explanation_quality': {
        'name': 'Explanation Quality',
        'high': 'Clear explanations in chat history',
        'low': 'Poor or missing explanations',
        'max_points': 10
    },
    'ai_usage_quality': {
        'name': 'AI Usage Quality',
        'high': 'Good use of AI assistance with learning focus',
        'low': 'Minimal AI usage or no learning evidence',
        'max_points': 10
    }
}

# Code-only rubric (when no chat history is available)
CODE_ONLY_RUBRIC = {
    'code_submission': {
        'name': 'Code Submission',
        'high': 'Code files present and well-structured',
        'low': 'No code files or very minimal',
        'max_points': 25
    },
    'completeness': {
        'name': 'Completeness',
        'high': 'All required components present',
        'low': 'Missing major components',
        'max_points': 25
    },
    'correctness': {
        'name': 'Code Correctness',
        'high': 'Code appears correct with good logic',
        'low': 'Code has obvious errors or issues',
        'max_points': 35
    },
    'code_quality': {
        'name': 'Code Quality',
        'high': 'Well-organized, good structure, proper comments',
        'low': 'Poorly organized, minimal structure',
        'max_points': 15
    }
}
# Add runtime criterion to code-only rubric
CODE_ONLY_RUBRIC['runtime'] = {
    'name': 'Runtime Behavior',
    'high': 'Program runs without crashing',
    'low': 'Program crashes or fails to run',
    'max_points': 10
}

# Student folder name patterns (adjust based on your submission format)
STUDENT_FOLDER_PATTERNS = [
    r'^[A-Za-z]+_\d+',           # e.g., "JohnDoe_12345"
    r'^\d+_[A-Za-z]+',           # e.g., "12345_JohnDoe"
    r'^[A-Za-z]+\s+\d+',         # e.g., "John Doe 12345"
    r'^student_',                # e.g., "student_12345"
    r'^[A-Z]{2,3}\d{6,8}',       # e.g., "ABC123456"
]


# =============================================================================
# GRADING FUNCTIONS
# =============================================================================

def detect_student_folders(base_path: Path) -> List[Path]:
    """
    Detect student submissions by grouping files by student name.
    Files are grouped by the prefix before the '#' symbol in the filename.
    .7z files are skipped as they cannot be processed for code extraction.
    """
    student_folders = {}
    skipped_files = []
    
    for item in base_path.iterdir():
        if item.is_file():
            file_name = item.name
            
            # Skip .7z files
            if file_name.lower().endswith('.7z'):
                skipped_files.append(item)
                continue
            
            # Extract student name from prefix before '#'
            if '#' in file_name:
                student_name = file_name.split('#')[0]
            else:
                # If no '#', use the filename without extension as student name
                student_name = item.stem
            
            # Group files by student name
            if student_name not in student_folders:
                student_folders[student_name] = []
            student_folders[student_name].append(item)
    
    # Report skipped files
    if skipped_files:
        print(f"  ⚠ Skipped {len(skipped_files)} .7z file(s):")
        for sf in skipped_files:
            print(f"    - {sf.name}")
    
    # Convert to list of student "folders" (groups of files)
    result = []
    for student_name, files in sorted(student_folders.items()):
        # Create a virtual folder path for each student
        student_path = base_path / student_name
        result.append((student_path, files))
    
    return result


def find_code_files(folder: Path, file_list: List[Path] = None) -> List[Path]:
    """Find all code files in a student folder or from a file list."""
    code_files = []
    
    if file_list:
        # Use provided file list (for file-based detection)
        for f in file_list:
            if f.is_file() and f'.{f.suffix[1:]}' in REQUIRED_CODE_EXTENSIONS:
                code_files.append(f)
    else:
        # Use folder search (for backward compatibility)
        for ext in REQUIRED_CODE_EXTENSIONS:
            code_files.extend(folder.rglob(f'*{ext}'))
    
    return code_files


def find_chat_history(folder: Path, file_list: List[Path] = None) -> List[Path]:
    """Find chat history files in a student folder or from a file list."""
    chat_files = []
    
    if file_list:
        # Use provided file list (for file-based detection)
        for f in file_list:
            if f.is_file() and f.name.lower() in ['chat.txt', 'conversation.txt', 'history.txt']:
                if f.suffix.lower() == '.txt':
                    chat_files.append(f)
    else:
        # Use folder search (for backward compatibility)
        for pattern in CHAT_HISTORY_PATTERNS:
            if '*' in pattern:
                # Handle wildcard patterns
                ext = pattern.replace('*', '')
                chat_files.extend(folder.rglob(f'*{ext}'))
            else:
                # Exact match
                for f in folder.rglob(pattern):
                    if f.is_file():
                        chat_files.append(f)
    
    return chat_files



def read_file_safe(file_path: Path, max_size_kb: int = 500) -> str:
    """
    Safely read a file as text (no execution).
    Limits file size to prevent memory issues.
    """
    try:
        # Check file size
        size_kb = file_path.stat().st_size / 1024
        if size_kb > max_size_kb:
            return f"[File too large: {size_kb:.1f}KB - truncated]"
        
        # Read as text (no execution)
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            return f.read()
    except Exception as e:
        return f"[Error reading file: {str(e)}]"


# -----------------------------------------------------------------------------
# Runtime smoke test function (optional)
# -----------------------------------------------------------------------------
def run_runtime_smoke_test(code_files: List[Path], timeout_seconds: int = RUNTIME_SMOKE_TEST_TIMEOUT_SECONDS) -> Dict:
    """
    Optionally launch submitted Python files for a short smoke test.

    IMPORTANT:
    - This executes student code, so only use it on trusted submissions.
    - It does not automatically judge gameplay quality.
    - For pygame programs, timing out after the window opens is considered a successful launch.
    """
    results = []

    python_files = [f for f in code_files if f.suffix.lower() == '.py']
    if not python_files:
        return {
            'enabled': True,
            'status': 'NO_PYTHON_FILES',
            'findings': ['No Python files available for runtime smoke test']
        }

    for code_file in python_files:
        try:
            proc = subprocess.run(
                [sys.executable, code_file.name],
                cwd=str(code_file.parent),
                capture_output=True,
                text=True,
                timeout=timeout_seconds
            )

            if proc.returncode == 0:
                results.append(f"✓ {code_file.name}: exited normally")
            else:
                error_preview = (proc.stderr or proc.stdout or '').strip().split('\n')[:2]
                error_text = ' | '.join(error_preview) if error_preview else 'No error output'
                results.append(f"✗ {code_file.name}: crashed or exited with code {proc.returncode}; {error_text}")

        except subprocess.TimeoutExpired:
            results.append(f"✓ {code_file.name}: launched and kept running for {timeout_seconds}s (likely pygame window opened)")
        except Exception as e:
            results.append(f"✗ {code_file.name}: runtime smoke test failed: {e}")

    any_success = any(item.startswith('✓') for item in results)
    any_failure = any(item.startswith('✗') for item in results)

    if any_success and not any_failure:
        status = 'PASS'
    elif any_success and any_failure:
        status = 'PARTIAL'
    else:
        status = 'FAIL'

    return {
        'enabled': True,
        'status': status,
        'findings': results
    }


def analyze_code_quality(code_files: List[Path], rubric: Dict = None) -> Tuple[int, str, List[str]]:
    """
    Analyze code files for quality indicators.
    Returns (score, status, findings).
    """
    if not code_files:
        return 0, 'LOW', ['No code files found']
    
    # Use RUBRIC by default, but allow override
    max_score = rubric['code_submission']['max_points'] if rubric else RUBRIC['code_submission']['max_points']
    
    total_lines = 0
    findings = []
    function_count = 0
    has_classes = False
    has_comments = False
    has_docstrings = False
    
    for code_file in code_files:
        content = read_file_safe(code_file)
        lines = content.split('\n')
        total_lines += len(lines)
        
        # Count functions
        function_count += len(re.findall(r'def\s+\w+\s*\(', content))
        
        # Check for classes
        if re.search(r'class\s+\w+', content):
            has_classes = True
        
        # Check for comments
        if re.search(r'#.*', content):
            has_comments = True
        
        # Check for docstrings
        if re.search(r'""".*?"""', content, re.DOTALL):
            has_docstrings = True
    
    findings.append(f"Total lines of code: {total_lines}")
    findings.append(f"Number of files: {len(code_files)}")
    findings.append(f"Number of functions: {function_count}")
    
    # Calculate score based on indicators
    score = 0
    
    # Base score for having code
    score += min(5, len(code_files) * 2)
    
    # Points for code structure - increased weight
    if function_count >= 3:
        score += 8
        findings.append("✓ Excellent modularity (multiple functions)")
    elif function_count >= 1:
        score += 5
        findings.append("✓ Functions detected")
    else:
        findings.append("✗ No functions detected")
    
    if has_classes:
        score += 5
        findings.append("✓ Classes detected")
    
    if has_comments:
        score += 3
        findings.append("✓ Comments present")
    
    if has_docstrings:
        score += 2
        findings.append("✓ Docstrings present")
    
    # Points for code volume - reduced weight
    if total_lines > 50:
        score += 1
        findings.append("✓ Substantial code volume")
    elif total_lines > 20:
        score += 1
        findings.append("✓ Moderate code volume")
    
    # Cap score
    score = min(score, max_score)
    
    # Determine status
    if score >= max_score * 0.8:
        status = 'HIGH'
    elif score >= max_score * 0.4:
        status = 'MEDIUM'
    else:
        status = 'LOW'
    
    return score, status, findings


def analyze_code_quality_only(code_files: List[Path]) -> Tuple[int, str, List[str]]:
    """
    Analyze code files specifically for code quality (structure, comments, etc.).
    Used for code-only rubric to get a separate code quality score.
    Returns (score, status, findings).
    """
    if not code_files:
        return 0, 'LOW', ['No code files found']
    
    max_score = CODE_ONLY_RUBRIC['code_quality']['max_points']
    
    findings = []
    total_lines = 0
    function_count = 0
    has_classes = False
    has_comments = False
    has_docstrings = False
    has_proper_naming = False
    has_error_handling = False
    
    for code_file in code_files:
        content = read_file_safe(code_file)
        lines = content.split('\n')
        total_lines += len(lines)
        
        # Count functions
        function_count += len(re.findall(r'def\s+\w+\s*\(', content))
        
        # Check for classes
        if re.search(r'class\s+\w+', content):
            has_classes = True
        
        # Check for comments
        if re.search(r'#.*', content):
            has_comments = True
        
        # Check for docstrings
        if re.search(r'""".*?"""', content, re.DOTALL):
            has_docstrings = True
        
        # Check for proper naming conventions
        if re.search(r'^[a-z][a-z0-9_]*\s*=|^[A-Z][A-Za-z0-9_]*\s*=', content, re.MULTILINE):
            has_proper_naming = True
        
        # Check for error handling
        if re.search(r'try:|except:|finally:', content):
            has_error_handling = True
    
    findings.append(f"Total lines of code: {total_lines}")
    findings.append(f"Number of files: {len(code_files)}")
    findings.append(f"Number of functions: {function_count}")
    
    # Calculate score
    score = 0
    
    # Reward modularity (multiple functions)
    if function_count >= 3:
        score += 4
        findings.append("✓ Good modularity (multiple functions)")
    elif function_count >= 1:
        score += 2
        findings.append("✓ Some modularity (functions present)")
    else:
        findings.append("✗ No functions detected")
    
    # Reward classes
    if has_classes:
        score += 3
        findings.append("✓ Classes detected")
    
    # Reward readability (comments)
    if has_comments:
        score += 3
        findings.append("✓ Comments present")
    
    # Reward documentation
    if has_docstrings:
        score += 2
        findings.append("✓ Docstrings present")
    
    # Reward proper naming
    if has_proper_naming:
        score += 2
        findings.append("✓ Proper naming conventions")
    
    # Reward error handling
    if has_error_handling:
        score += 2
        findings.append("✓ Error handling present")
    
    # Cap score
    score = min(score, max_score)
    
    # Determine status
    if score >= max_score * 0.8:
        status = 'HIGH'
    elif score >= max_score * 0.4:
        status = 'MEDIUM'
    else:
        status = 'LOW'
    
    return score, status, findings


def analyze_chat_history(chat_files: List[Path]) -> Tuple[int, str, List[str]]:
    """
    Analyze chat history files for quality.
    Returns (score, status, findings).
    """
    if not chat_files:
        return 0, 'LOW', ['No chat history found']
    
    total_chars = 0
    findings = []
    has_questions = False
    has_explanations = False
    has_code_discussion = False
    
    for chat_file in chat_files:
        content = read_file_safe(chat_file)
        total_chars += len(content)
        
        # Check for questions (learning attempt)
        if re.search(r'\?|how|why|what|explain', content, re.IGNORECASE):
            has_questions = True
        
        # Check for explanations
        if re.search(r'because|therefore|so that|this means', content, re.IGNORECASE):
            has_explanations = True
        
        # Check for code discussion
        if re.search(r'code|function|class|variable|algorithm', content, re.IGNORECASE):
            has_code_discussion = True
    
    findings.append(f"Total chat characters: {total_chars}")
    findings.append(f"Number of chat files: {len(chat_files)}")
    
    # Calculate score
    score = 0
    max_score = RUBRIC['chat_history']['max_points']
    
    # Base score for having chat
    score += min(5, len(chat_files) * 3)
    
    # Points for content quality
    if has_questions:
        score += 4
        findings.append("✓ Questions present (learning)")
    if has_explanations:
        score += 4
        findings.append("✓ Explanations present")
    if has_code_discussion:
        score += 2
        findings.append("✓ Code discussion present")
    
    # Points for volume
    if total_chars > 1000:
        score += 5
        findings.append("✓ Substantial chat history")
    elif total_chars > 300:
        score += 3
        findings.append("✓ Moderate chat history")
    
    # Cap score
    score = min(score, max_score)
    
    # Determine status
    if score >= max_score * 0.8:
        status = 'HIGH'
    elif score >= max_score * 0.4:
        status = 'MEDIUM'
    else:
        status = 'LOW'
    
    return score, status, findings


def analyze_completeness(code_files: List[Path], chat_files: List[Path], use_code_only: bool = False) -> Tuple[int, str, List[str]]:
    """
    Analyze overall submission completeness.
    Returns (score, status, findings).
    """
    findings = []
    
    if use_code_only:
        # Code-only mode: only check for code files
        max_score = CODE_ONLY_RUBRIC['completeness']['max_points']
        
        if code_files:
            score = max_score
            findings.append(f"✓ Code files: {len(code_files)}")
            findings.append("✓ Submission complete (code-only mode)")
        else:
            score = 0
            findings.append("✗ No code files")
    else:
        # Full rubric mode: check both code and chat
        max_score = RUBRIC['completeness']['max_points']
        score = 0
        
        # Check code files
        if code_files:
            score += 10
            findings.append(f"✓ Code files: {len(code_files)}")
        else:
            findings.append("✗ No code files")
        
        # Check chat history
        if chat_files:
            score += 10
            findings.append(f"✓ Chat history: {len(chat_files)}")
        else:
            findings.append("✗ No chat history")
    
    # Determine status
    if score >= max_score * 0.8:
        status = 'HIGH'
    elif score >= max_score * 0.4:
        status = 'MEDIUM'
    else:
        status = 'LOW'
    
    return score, status, findings


def analyze_correctness(code_files: List[Path], rubric: Dict = None) -> Tuple[int, str, List[str]]:
    """
    Analyze code for correctness (static analysis only - no execution).
    Returns (score, status, findings).
    """
    if not code_files:
        return 0, 'LOW', ['No code to analyze']
    
    findings = []
    error_count = 0
    warning_count = 0
    has_game_loop = False
    has_pygame = False
    has_rendering = False
    has_entry_point = False
    
    for code_file in code_files:
        content = read_file_safe(code_file)
        
        # Check for common syntax issues (static analysis)
        # Check for unclosed brackets/braces
        open_braces = content.count('{')
        close_braces = content.count('}')
        open_brackets = content.count('[')
        close_brackets = content.count(']')
        open_parens = content.count('(')
        close_parens = content.count(')')
        
        if abs(open_braces - close_braces) > 2:
            warning_count += 1
            findings.append(f"Warning: Possible unclosed braces in {code_file.name}")
        
        if abs(open_brackets - close_brackets) > 2:
            warning_count += 1
            findings.append(f"Warning: Possible unclosed brackets in {code_file.name}")
        
        if abs(open_parens - close_parens) > 2:
            warning_count += 1
            findings.append(f"Warning: Possible unclosed parentheses in {code_file.name}")
        
        # Check for TODO/FIXME (indicates incomplete work) - NO LONGER DEDUCT POINTS
        if re.search(r'TODO|FIXME|XXX|HACK', content):
            findings.append(f"Note: TODO/FIXME markers in {code_file.name}")
        
        # Check for core program structure
        # Game loop (while loop)
        if re.search(r'while\s+.*:', content):
            has_game_loop = True
        
        # Pygame usage
        if re.search(r'import\s+pygame', content):
            has_pygame = True
        
        # Rendering/update logic
        if re.search(r'pygame\.display|update\(\)|draw\(\)', content):
            has_rendering = True
        
        # Entry point
        content_lower = content.lower()
        if 'def main' in content_lower or 'if __name__' in content_lower or 'public static void main' in content_lower:
            has_entry_point = True
    
    # Calculate score
    score = 0
    max_score = rubric['correctness']['max_points'] if rubric else RUBRIC['correctness']['max_points']
    
    # Base score for having code
    score += 10
    
    # Deduct for warnings (syntax issues)
    score -= warning_count * 2
    score -= error_count * 5
    
    # Add for good patterns
    if warning_count == 0:
        score += 5
        findings.append("✓ No obvious syntax issues")
    
    # Reward for core program structure
    if has_game_loop:
        score += 8
        findings.append("✓ Game loop detected")
    else:
        score -= 10
        findings.append("✗ No game loop detected")
    
    if has_pygame:
        score += 8
        findings.append("✓ Pygame usage detected")
    else:
        score -= 10
        findings.append("✗ No pygame usage detected")
    
    if has_rendering:
        score += 5
        findings.append("✓ Rendering/update logic detected")
    else:
        score -= 5
        findings.append("✗ No rendering/update logic detected")
    
    if has_entry_point:
        score += 5
        findings.append("✓ Entry point detected")
    else:
        score -= 5
        findings.append("✗ No entry point detected")
    
    # Cap score
    score = max(0, min(score, max_score))
    
    # Determine status
    if score >= max_score * 0.8:
        status = 'HIGH'
    elif score >= max_score * 0.4:
        status = 'MEDIUM'
    else:
        status = 'LOW'
    
    return score, status, findings


def analyze_explanation_quality(chat_files: List[Path]) -> Tuple[int, str, List[str]]:
    """
    Analyze explanation quality in chat history.
    Returns (score, status, findings).
    """
    if not chat_files:
        return 0, 'LOW', ['No chat history to analyze']
    
    total_content = ''
    for chat_file in chat_files:
        total_content += read_file_safe(chat_file) + '\n'
    
    findings = []
    score = 0
    max_score = RUBRIC['explanation_quality']['max_points']
    
    # Check for explanation indicators
    explanation_patterns = [
        (r'\b(because|therefore|thus|hence)\b', 'Cause/effect reasoning'),
        (r'\b(this means|in other words|essentially)\b', 'Clarification'),
        (r'\b(first|second|third|finally|additionally)\b', 'Structured explanation'),
        (r'\b(for example|for instance|specifically)\b', 'Examples given'),
    ]
    
    for pattern, description in explanation_patterns:
        if re.search(pattern, total_content, re.IGNORECASE):
            score += 2
            findings.append(f"✓ {description}")
    
    # Check for technical terms usage
    technical_terms = len(re.findall(r'\b(function|class|method|variable|algorithm|data|loop|condition)\b', total_content, re.IGNORECASE))
    if technical_terms > 5:
        score += 3
        findings.append("✓ Good use of technical terminology")
    
    # Cap score
    score = min(score, max_score)
    
    # Determine status
    if score >= max_score * 0.8:
        status = 'HIGH'
    elif score >= max_score * 0.4:
        status = 'MEDIUM'
    else:
        status = 'LOW'
    
    return score, status, findings


def analyze_ai_usage_quality(chat_files: List[Path]) -> Tuple[int, str, List[str]]:
    """
    Analyze AI usage quality in chat history.
    Returns (score, status, findings).
    """
    if not chat_files:
        return 0, 'LOW', ['No chat history to analyze']
    
    total_content = ''
    for chat_file in chat_files:
        total_content += read_file_safe(chat_file) + '\n'
    
    findings = []
    score = 0
    max_score = RUBRIC['ai_usage_quality']['max_points']
    
    # Check for learning indicators
    learning_patterns = [
        (r'\b(understand|learn|explain|help me)\b', 'Learning intent'),
        (r'\b(why|how|what if|could you explain)\b', 'Questions asked'),
        (r'\b(thanks|thank you|great|perfect)\b', 'Acknowledgment'),
    ]
    
    for pattern, description in learning_patterns:
        if re.search(pattern, total_content, re.IGNORECASE):
            score += 2
            findings.append(f"✓ {description}")
    
    # Check for iterative learning (follow-up questions)
    follow_ups = len(re.findall(r'(but|however|also|another|more)', total_content, re.IGNORECASE))
    if follow_ups > 3:
        score += 2
        findings.append("✓ Iterative learning detected")
    
    # Cap score
    score = min(score, max_score)
    
    # Determine status
    if score >= max_score * 0.8:
        status = 'HIGH'
    elif score >= max_score * 0.4:
        status = 'MEDIUM'
    else:
        status = 'LOW'
    
    return score, status, findings


def grade_student(folder: Path, file_list: List[Path] = None, use_code_only: bool = False) -> Dict:
    """
    Grade a single student's submission.
    Returns a dictionary with all grading information.
    """
    # Use the appropriate rubric
    rubric = CODE_ONLY_RUBRIC if use_code_only else RUBRIC

    # Find files
    code_files = find_code_files(folder, file_list)
    chat_files = find_chat_history(folder, file_list)

    runtime_check = {'enabled': False, 'status': 'NOT_RUN', 'findings': ['Runtime smoke test not enabled']}
    if os.environ.get('RUN_SMOKE_TEST', '').lower() in {'1', 'true', 'yes'}:
        runtime_check = run_runtime_smoke_test(code_files)

    if use_code_only:
        # Code-only grading: use CODE_ONLY_RUBRIC weights
        code_score, code_status, code_findings = analyze_code_quality(code_files, CODE_ONLY_RUBRIC)
        completeness_score, completeness_status, completeness_findings = analyze_completeness(code_files, chat_files, use_code_only=True)
        correctness_score, correctness_status, correctness_findings = analyze_correctness(code_files, CODE_ONLY_RUBRIC)
        code_quality_score, code_quality_status, code_quality_findings = analyze_code_quality_only(code_files)

        # Runtime scoring
        runtime_score = 0
        runtime_status = 'N/A'
        runtime_findings = []

        if runtime_check['enabled']:
            if runtime_check['status'] == 'PASS':
                runtime_score = 10
                runtime_status = 'HIGH'
                runtime_findings.append("✓ Program runs without crashing")
            elif runtime_check['status'] == 'PARTIAL':
                runtime_score = 5
                runtime_status = 'MEDIUM'
                runtime_findings.append("⚠ Some files failed during runtime")
            else:
                runtime_score = 0
                runtime_status = 'LOW'
                runtime_findings.append("✗ Program crashes or fails to run")

        total_score = code_score + completeness_score + correctness_score + code_quality_score + runtime_score
        max_total = sum(r['max_points'] for r in CODE_ONLY_RUBRIC.values())

        return {
            'student_name': folder.name,
            'folder_path': str(folder),
            'code_files': [str(f) for f in code_files],
            'chat_files': [str(f) for f in chat_files],
            'use_code_only': use_code_only,
            'runtime_check': runtime_check,
            'code_submission': {
                'score': code_score,
                'status': code_status,
                'findings': code_findings
            },
            'completeness': {
                'score': completeness_score,
                'status': completeness_status,
                'findings': completeness_findings
            },
            'correctness': {
                'score': correctness_score,
                'status': correctness_status,
                'findings': correctness_findings
            },
            'code_quality': {
                'score': code_quality_score,
                'status': code_quality_status,
                'findings': code_quality_findings
            },
            'runtime': {
                'score': runtime_score,
                'status': runtime_status,
                'findings': runtime_findings
            },
            'chat_history': {
                'score': 0,
                'status': 'N/A',
                'findings': ['Not applicable in code-only mode']
            },
            'explanation_quality': {
                'score': 0,
                'status': 'N/A',
                'findings': ['Not applicable in code-only mode']
            },
            'ai_usage_quality': {
                'score': 0,
                'status': 'N/A',
                'findings': ['Not applicable in code-only mode']
            },
            'total_score': total_score,
            'max_total': max_total,
            'percentage': round(total_score / max_total * 100, 1)
        }
    else:
        # Full rubric grading (with chat history)
        code_score, code_status, code_findings = analyze_code_quality(code_files, RUBRIC)
        chat_score, chat_status, chat_findings = analyze_chat_history(chat_files)
        completeness_score, completeness_status, completeness_findings = analyze_completeness(code_files, chat_files, use_code_only=False)
        correctness_score, correctness_status, correctness_findings = analyze_correctness(code_files, RUBRIC)
        explanation_score, explanation_status, explanation_findings = analyze_explanation_quality(chat_files)
        ai_usage_score, ai_usage_status, ai_usage_findings = analyze_ai_usage_quality(chat_files)

        # Calculate total using RUBRIC weights
        total_score = (code_score + chat_score + completeness_score +
                       correctness_score + explanation_score + ai_usage_score)
        max_total = sum(r['max_points'] for r in RUBRIC.values())

        return {
            'student_name': folder.name,
            'folder_path': str(folder),
            'code_files': [str(f) for f in code_files],
            'chat_files': [str(f) for f in chat_files],
            'use_code_only': use_code_only,
            'runtime_check': runtime_check,
            'code_submission': {
                'score': code_score,
                'status': code_status,
                'findings': code_findings
            },
            'chat_history': {
                'score': chat_score,
                'status': chat_status,
                'findings': chat_findings
            },
            'completeness': {
                'score': completeness_score,
                'status': completeness_status,
                'findings': completeness_findings
            },
            'correctness': {
                'score': correctness_score,
                'status': correctness_status,
                'findings': correctness_findings
            },
            'explanation_quality': {
                'score': explanation_score,
                'status': explanation_status,
                'findings': explanation_findings
            },
            'ai_usage_quality': {
                'score': ai_usage_score,
                'status': ai_usage_status,
                'findings': ai_usage_findings
            },
            'code_quality': {
                'score': 0,
                'status': 'N/A',
                'findings': ['Not applicable in full rubric mode']
            },
            'total_score': total_score,
            'max_total': max_total,
            'percentage': round(total_score / max_total * 100, 1)
        }


# =============================================================================
# OUTPUT FUNCTIONS
# =============================================================================

def create_excel_output(grades: List[Dict], output_path: str):
    """Create Excel file with grading results."""
    wb = Workbook()
    
    # Determine which rubric was used
    use_code_only = grades[0].get('use_code_only', False) if grades else False
    rubric = CODE_ONLY_RUBRIC if use_code_only else RUBRIC
    
    # =========================================================================
    # Sheet 1: Student Grades
    # =========================================================================
    ws_grades = wb.active
    ws_grades.title = "Student Grades"
    
    # Headers - adjust based on rubric
    if use_code_only:
        headers = [
            'Student Name',
            'Code Submission Score',
            'Completeness Score',
            'Correctness Score',
            'Code Quality Score',
            'Runtime Score',
            'Total Score',
            'Max Score',
            'Percentage',
            'Grade'
        ]
    else:
        headers = [
            'Student Name',
            'Code Submission Score',
            'Chat History Score',
            'Completeness Score',
            'Correctness Score',
            'Explanation Score',
            'AI Usage Score',
            'Total Score',
            'Max Score',
            'Percentage',
            'Grade'
        ]
    
    for col, header in enumerate(headers, 1):
        cell = ws_grades.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    for row, grade in enumerate(grades, 2):
        # Determine grade
        pct = grade['percentage']
        if pct >= 90:
            letter_grade = 'A'
        elif pct >= 80:
            letter_grade = 'B'
        elif pct >= 70:
            letter_grade = 'C'
        elif pct >= 60:
            letter_grade = 'D'
        else:
            letter_grade = 'F'
        
        if use_code_only:
            row_data = [
                grade['student_name'],
                f"{grade['code_submission']['score']} / {CODE_ONLY_RUBRIC['code_submission']['max_points']}",
                f"{grade['completeness']['score']} / {CODE_ONLY_RUBRIC['completeness']['max_points']}",
                f"{grade['correctness']['score']} / {CODE_ONLY_RUBRIC['correctness']['max_points']}",
                f"{grade['code_quality']['score']} / {CODE_ONLY_RUBRIC['code_quality']['max_points']}",
                f"{grade.get('runtime', {}).get('score', 0)} / 10",
                grade['total_score'],
                grade['max_total'],
                f"{grade['percentage']}%",
                letter_grade
            ]
        else:
            row_data = [
                grade['student_name'],
                f"{grade['code_submission']['score']} / {RUBRIC['code_submission']['max_points']}",
                f"{grade['chat_history']['score']} / {RUBRIC['chat_history']['max_points']}",
                f"{grade['completeness']['score']} / {RUBRIC['completeness']['max_points']}",
                f"{grade['correctness']['score']} / {RUBRIC['correctness']['max_points']}",
                f"{grade['explanation_quality']['score']} / {RUBRIC['explanation_quality']['max_points']}",
                f"{grade['ai_usage_quality']['score']} / {RUBRIC['ai_usage_quality']['max_points']}",
                grade['total_score'],
                grade['max_total'],
                f"{grade['percentage']}%",
                letter_grade
            ]
        
        for col, value in enumerate(row_data, 1):
            ws_grades.cell(row=row, column=col).value = value
    
    # Adjust column widths
    ws_grades.column_dimensions['A'].width = 20
    if use_code_only:
        for col in 'BCDE':
            ws_grades.column_dimensions[col].width = 18
        for col in 'FG':
            ws_grades.column_dimensions[col].width = 10
        ws_grades.column_dimensions['H'].width = 12
        ws_grades.column_dimensions['I'].width = 8
    else:
        for col in 'BCDEFGH':
            ws_grades.column_dimensions[col].width = 16
        for col in 'IJ':
            ws_grades.column_dimensions[col].width = 10
        ws_grades.column_dimensions['K'].width = 8
    
    # =========================================================================
    # Sheet 2: Rubric
    # =========================================================================
    ws_rubric = wb.create_sheet("Rubric")
    
    # Rubric headers
    rubric_headers = ['Criterion', 'Description', 'Max Points', 'High Performance', 'Low Performance']
    for col, header in enumerate(rubric_headers, 1):
        cell = ws_rubric.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    # Rubric data - use the appropriate rubric
    for row, (key, rubric_item) in enumerate(rubric.items(), 2):
        ws_rubric.cell(row=row, column=1).value = rubric_item['name']
        ws_rubric.cell(row=row, column=2).value = key.replace('_', ' ').title()
        ws_rubric.cell(row=row, column=3).value = rubric_item['max_points']
        ws_rubric.cell(row=row, column=4).value = rubric_item['high']
        ws_rubric.cell(row=row, column=5).value = rubric_item['low']
    
    # Adjust column widths
    ws_rubric.column_dimensions['A'].width = 20
    ws_rubric.column_dimensions['B'].width = 20
    ws_rubric.column_dimensions['C'].width = 12
    ws_rubric.column_dimensions['D'].width = 40
    ws_rubric.column_dimensions['E'].width = 40
    
    # =========================================================================
    # Sheet 3: Detailed Scores
    # =========================================================================
    ws_detail = wb.create_sheet("Detailed Scores")
    
    # Headers - adjust based on rubric
    if use_code_only:
        detail_headers = ['Student Name', 'Code Score', 'Completeness', 
                         'Correctness', 'Code Quality', 'Runtime', 'Total']
    else:
        detail_headers = ['Student Name', 'Code Score', 'Chat Score', 'Completeness', 
                         'Correctness', 'Explanation', 'AI Usage', 'Total']
    for col, header in enumerate(detail_headers, 1):
        cell = ws_detail.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    # Data rows
    for row, grade in enumerate(grades, 2):
        if use_code_only:
            row_data = [
                grade['student_name'],
                grade['code_submission']['score'],
                grade['completeness']['score'],
                grade['correctness']['score'],
                grade['code_quality']['score'],
                grade.get('runtime', {}).get('score', 0),
                grade['total_score']
            ]
        else:
            row_data = [
                grade['student_name'],
                grade['code_submission']['score'],
                grade['chat_history']['score'],
                grade['completeness']['score'],
                grade['correctness']['score'],
                grade['explanation_quality']['score'],
                grade['ai_usage_quality']['score'],
                grade['total_score']
            ]
        
        for col, value in enumerate(row_data, 1):
            ws_detail.cell(row=row, column=col).value = value
    
    # Adjust column widths
    ws_detail.column_dimensions['A'].width = 20
    if use_code_only:
        for col in 'BCDEF':
            ws_detail.column_dimensions[col].width = 12
    else:
        for col in 'BCDEFGH':
            ws_detail.column_dimensions[col].width = 12
    
    # Save workbook
    wb.save(output_path)
    print(f"✓ Excel file created: {output_path}")


def create_grading_log(grades: List[Dict], output_path: str):
    """Create grading log text file."""
    # Determine which rubric was used
    use_code_only = grades[0].get('use_code_only', False) if grades else False
    rubric = CODE_ONLY_RUBRIC if use_code_only else RUBRIC
    
    with open(output_path, 'w') as f:
        f.write("=" * 80 + "\n")
        f.write("STUDENT HOMEWORK GRADING LOG\n")
        f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        
        # Add rubric info
        if use_code_only:
            f.write("Rubric: CODE-ONLY (scores rescaled to 100)\n")
        else:
            f.write("Rubric: FULL (code + chat history)\n")
        
        f.write("=" * 80 + "\n\n")
        
        for grade in grades:
            f.write("-" * 80 + "\n")
            f.write(f"STUDENT: {grade['student_name']}\n")
            f.write(f"Folder: {grade['folder_path']}\n")
            f.write("-" * 80 + "\n")
            
            # Files submitted
            f.write("\n📁 FILES SUBMITTED:\n")
            if grade['code_files']:
                f.write("  Code Files:\n")
                for cf in grade['code_files']:
                    f.write(f"    - {os.path.basename(cf)}\n")
            else:
                f.write("  Code Files: None\n")
            
            if grade['chat_files']:
                f.write("  Chat History:\n")
                for cf in grade['chat_files']:
                    f.write(f"    - {os.path.basename(cf)}\n")
            else:
                f.write("  Chat History: None\n")

            runtime_check = grade.get('runtime_check', {'enabled': False, 'status': 'NOT_RUN', 'findings': []})
            f.write("\n🧪 RUNTIME SMOKE TEST:\n")
            f.write(f"  Status: {runtime_check.get('status', 'NOT_RUN')}\n")
            for finding in runtime_check.get('findings', []):
                f.write(f"    • {finding}\n")
            
            # Grading details for each criterion
            f.write("\n📊 GRADING DETAILS:\n")
            
            if use_code_only:
                criteria = [
                    ('code_submission', 'Code Submission'),
                    ('completeness', 'Completeness'),
                    ('correctness', 'Code Correctness'),
                    ('code_quality', 'Code Quality'),
                    ('runtime', 'Runtime Behavior'),
                ]
            else:
                criteria = [
                    ('code_submission', 'Code Submission'),
                    ('chat_history', 'Chat History'),
                    ('completeness', 'Completeness'),
                    ('correctness', 'Code Correctness'),
                    ('explanation_quality', 'Explanation Quality'),
                    ('ai_usage_quality', 'AI Usage Quality')
                ]
            
            for key, name in criteria:
                criterion = grade[key]
                max_pts = rubric[key]['max_points'] if key in rubric else 0
                f.write(f"\n  [{criterion['status']}] {name}: {criterion['score']}/{max_pts} points\n")
                for finding in criterion['findings']:
                    f.write(f"    • {finding}\n")
            
            # Total
            f.write("\n" + "=" * 40 + "\n")
            f.write(f"TOTAL SCORE: {grade['total_score']}/{grade['max_total']} ({grade['percentage']}%)\n")
            f.write("=" * 40 + "\n\n")
        
        # Summary
        f.write("=" * 80 + "\n")
        f.write("SUMMARY\n")
        f.write("=" * 80 + "\n")
        f.write(f"Total Students: {len(grades)}\n")
        
        if grades:
            avg_percentage = sum(g['percentage'] for g in grades) / len(grades)
            f.write(f"Average Score: {avg_percentage:.1f}%\n")
            
            high_count = sum(1 for g in grades if g['percentage'] >= 80)
            medium_count = sum(1 for g in grades if 60 <= g['percentage'] < 80)
            low_count = sum(1 for g in grades if g['percentage'] < 60)
            
            f.write(f"High Performance (≥80%): {high_count}\n")
            f.write(f"Medium Performance (60-79%): {medium_count}\n")
            f.write(f"Low Performance (<60%): {low_count}\n")
    
    print(f"✓ Grading log created: {output_path}")


# =============================================================================
# MAIN FUNCTION
# =============================================================================

def main():
    """Main function to run the grader."""
    print("=" * 60)
    print("Student Homework Submission Grader")
    print("=" * 60)
    
    # Get ZIP file path
    if len(sys.argv) > 1:
        zip_path = sys.argv[1]
    else:
        # Look for ZIP files in current directory
        zip_files = list(Path('.').glob('*.zip'))
        if not zip_files:
            print("\n❌ No ZIP file found!")
            print("\nUsage: python homework_grader.py <path_to_zip_file>")
            print("\nOr place your ZIP file in the current directory and run:")
            print("  python homework_grader.py")
            sys.exit(1)
        elif len(zip_files) == 1:
            zip_path = str(zip_files[0])
        else:
            print("\nMultiple ZIP files found. Please specify one:")
            for i, zf in enumerate(zip_files, 1):
                print(f"  {i}. {zf.name}")
            sys.exit(1)
    
    print(f"\n📦 Processing: {zip_path}")
    
    # Create temporary extraction directory
    extract_dir = Path('_extracted_submissions')
    if extract_dir.exists():
        shutil.rmtree(extract_dir)
    extract_dir.mkdir()
    
    try:
        # Unzip submissions
        print("\n📂 Extracting submissions...")
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            zip_ref.extractall(extract_dir)
        print(f"✓ Extracted to: {extract_dir}")
        
        # Detect student submissions (now file-based)
        print("\n👥 Detecting student submissions...")
        student_groups = detect_student_folders(extract_dir)
        
        if not student_groups:
            print("❌ No student submissions detected!")
            print("\nExpected file structure:")
            print("  - Files should be named with student ID# pattern")
            print("  - Code files: .py, .java, .js, .ts, .cpp, .c, .go, .rs")
            print("  - .7z files will be skipped")
            sys.exit(1)
        
        print(f"✓ Found {len(student_groups)} student submissions")
        
        # Check if any chat history files exist across all students
        print("\n🔍 Checking for chat history files...")
        has_chat_history = False
        for student_path, files in student_groups:
            chat_files = find_chat_history(student_path, files)
            if chat_files:
                has_chat_history = True
                break
        
        # Determine which rubric to use
        use_code_only = not has_chat_history
        if use_code_only:
            print("⚠ No chat history files found for any student!")
            print("📋 Switching to CODE-ONLY rubric (scores rescaled to 100)")
            print(f"   Rubric: {', '.join(CODE_ONLY_RUBRIC.keys())}")
        else:
            print("✓ Chat history files detected - using full rubric")
        
        # Grade each student
        print("\n📝 Grading submissions...")
        grades = []
        
        for i, (student_path, files) in enumerate(student_groups, 1):
            student_name = student_path.name
            print(f"  [{i}/{len(student_groups)}] Grading: {student_name} ({len(files)} file(s))")
            grade = grade_student(student_path, files, use_code_only)
            grades.append(grade)
        
        # Sort by student name
        grades.sort(key=lambda x: x['student_name'])
        
        # Create outputs
        print("\n📊 Creating output files...")
        
        base_name = Path(zip_path).stem
        excel_path = f"{base_name}_grades.xlsx"
        log_path = f"{base_name}_grading_log.txt"
        
        create_excel_output(grades, excel_path)
        create_grading_log(grades, log_path)
        
        # Print summary
        print("\n" + "=" * 60)
        print("GRADING COMPLETE")
        print("=" * 60)
        print(f"\n📊 Results: {excel_path}")
        print(f"📝 Log: {log_path}")
        print(f"\nTotal Students: {len(grades)}")
        
        if grades:
            avg = sum(g['percentage'] for g in grades) / len(grades)
            print(f"Average Score: {avg:.1f}%")
            
            # Grade distribution
            a_count = sum(1 for g in grades if g['percentage'] >= 90)
            b_count = sum(1 for g in grades if 80 <= g['percentage'] < 90)
            c_count = sum(1 for g in grades if 70 <= g['percentage'] < 80)
            d_count = sum(1 for g in grades if 60 <= g['percentage'] < 70)
            f_count = sum(1 for g in grades if g['percentage'] < 60)
            
            print(f"\nGrade Distribution:")
            print(f"  A: {a_count} | B: {b_count} | C: {c_count} | D: {d_count} | F: {f_count}")
        
    finally:
        # Cleanup
        if extract_dir.exists():
            print(f"\n🧹 Cleaning up extraction folder...")
            # Keep the folder for inspection if needed
            # shutil.rmtree(extract_dir)
            print(f"  (Extraction folder kept: {extract_dir})")


if __name__ == '__main__':
    main()